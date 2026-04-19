"""Excel workbook utilities for notebooks and scripts.

NotebookWorkbook  — open or create an .xlsx, write DataFrames to named sheets,
                    preserve any sheets not being overwritten.
NotebookOutput    — resolves the local output/ folder next to the notebook and
                    provides helpers for Excel workbooks and plot paths.

Output convention
-----------------
Every notebook saves to its OWN local output/ subfolder, not the global
project-level output/.  Use NotebookOutput at the top of every notebook::

    from tools.excel import NotebookOutput
    OUT = NotebookOutput()          # → <notebook_dir>/output/

    wb  = OUT.excel("analysis.xlsx")   # NotebookWorkbook
    fig.savefig(OUT / "chart.png")     # Path
    OUT.path("data.csv")               # Path for any file type
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class NotebookWorkbook:
    """Write DataFrames to sheets in an Excel workbook.

    Opens an existing workbook or creates a new one.  Each call to
    ``write()`` replaces that sheet (if it exists) without touching others.
    Call ``save()`` once at the end.

    Usage::

        from pathlib import Path
        from tools.excel import NotebookWorkbook

        wb = NotebookWorkbook(Path("output/my_model.xlsx"))
        wb.write("Yields", yields_df)
        wb.write("Term Premium", tp_df)
        wb.save()
    """

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            self._wb = load_workbook(self.path)
        else:
            self._wb = Workbook()
            self._wb.remove(self._wb.active)

    def write(self, sheet_name: str, df: pd.DataFrame, description_row: bool = False) -> None:
        """Write *df* to *sheet_name*, replacing it if it already exists."""
        if sheet_name in self._wb.sheetnames:
            del self._wb[sheet_name]
        ws = self._wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        # Widen first column if it looks like dates
        ws.column_dimensions["A"].width = 14

    def save(self) -> Path:
        """Save the workbook and return the path."""
        self._wb.save(self.path)
        return self.path

    def __repr__(self) -> str:
        sheets = self._wb.sheetnames
        return f"NotebookWorkbook({self.path}, sheets={sheets})"


class NotebookOutput:
    """Resolves and manages the local output/ folder for a notebook.

    By default uses the current working directory (which equals the notebook's
    directory when Jupyter is launched from that folder).  Pass *base* to
    override — useful in scripts or when CWD differs from the notebook location.

    Usage::

        from tools.excel import NotebookOutput

        OUT = NotebookOutput()                   # <cwd>/output/
        OUT = NotebookOutput("models/my_model")  # explicit base

        # Excel workbook
        wb = OUT.excel("results.xlsx")
        wb.write("Sheet1", df)
        wb.save()

        # Plot path (use with savefig)
        fig.savefig(OUT / "chart.png", dpi=150, bbox_inches="tight")

        # Any file path
        df.to_csv(OUT.path("data.csv"), index=False)
    """

    def __init__(self, base: str | Path | None = None):
        self.base = Path(base).resolve() if base else Path().resolve()
        self.output_dir = self.base / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def path(self, filename: str) -> Path:
        """Return an absolute path inside the local output/ folder."""
        return self.output_dir / filename

    def excel(self, filename: str) -> NotebookWorkbook:
        """Open (or create) a NotebookWorkbook inside the local output/ folder."""
        if not filename.endswith(".xlsx"):
            filename = filename + ".xlsx"
        return NotebookWorkbook(self.path(filename))

    def __truediv__(self, filename: str) -> Path:
        """Support ``OUT / "chart.png"`` syntax."""
        return self.path(filename)

    def __repr__(self) -> str:
        return f"NotebookOutput({self.output_dir})"
